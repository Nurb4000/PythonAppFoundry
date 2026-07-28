#!/usr/bin/env python3
"""Tests for multi-format export (CSV, JSON, XLSX, PDF).

Run with: python test_export.py
Requires: openpyxl, reportlab
"""
import sys
import os
import io
import json
import csv
from flask import Flask

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '.'))

from app import create_app, db as _db
from app.services.exporters import (
    _export_json, _export_xlsx, _export_pdf,
    EXPORT_FORMATS, _serialize_value, _row_to_dict, HAS_OPENPYXL, HAS_REPORTLAB,
)
from app.services.admin_utils import _export_csv


PASS = '\033[32m\u2713\033[0m'
FAIL = '\033[31m\u2717\033[0m'

passed = 0
failed = 0


def check(condition, name):
    global passed, failed
    if condition:
        print(f"  {PASS} {name}")
        passed += 1
    else:
        print(f"  {FAIL} {name}")
        failed += 1


# ---------------------------------------------------------------------------
# Unit tests for the exporter functions (no DB needed)
# ---------------------------------------------------------------------------

class FakeRow:
    """Minimal stand-in for a SQLAlchemy model row."""
    def __init__(self, **kwargs):
        self.__dict__.update(kwargs)


def test_serialize_value():
    print("\n-- _serialize_value --")
    from datetime import datetime, date
    check(_serialize_value(None) is None, 'None -> None')
    check(_serialize_value(42) == 42, 'int passes through')
    check(_serialize_value('hi') == 'hi', 'str passes through')
    check(isinstance(_serialize_value(datetime.now()), str), 'datetime -> isostring')
    check(isinstance(_serialize_value(date.today()), str), 'date -> isostring')


def test_row_to_dict():
    print("\n-- _row_to_dict --")
    r = FakeRow(id=1, name='test', module=FakeRow(name='mymod'))
    d = _row_to_dict(r, ['id', 'name'], has_module=True)
    check(d['module'] == 'mymod', 'module column added')
    check(d['id'] == 1, 'id serialized')
    check(d['name'] == 'test', 'name serialized')


def test_export_json():
    print("\n-- _export_json --")
    rows = [FakeRow(id=1, name='alice'), FakeRow(id=2, name='bob')]
    resp = _export_json('users', ['id', 'name'], rows, False)
    check(resp.mimetype == 'application/json', 'content-type is application/json')
    check('attachment' in resp.headers.get('Content-Disposition', ''), 'has Content-Disposition')
    data = json.loads(resp.data.decode())
    check(data['name'] == 'users', 'top-level name field')
    check(data['columns'] == ['id', 'name'], 'columns preserved')
    check(len(data['rows']) == 2, 'two rows exported')
    check(data['rows'][0]['name'] == 'alice', 'row data correct')


def test_export_xlsx():
    print("\n-- _export_xlsx --")
    if not HAS_OPENPYXL:
        print("  SKIPPED (openpyxl not installed)")
        return
    rows = [FakeRow(id=1, name='alice'), FakeRow(id=2, name='bob')]
    resp = _export_xlsx('users', ['id', 'name'], rows, False)
    check(resp.status_code == 200, f'status 200 (got {resp.status_code})')
    check(resp.mimetype == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
          'correct xlsx content-type')
    buf = io.BytesIO(resp.data)
    import openpyxl
    wb = openpyxl.load_workbook(buf)
    ws = wb.active
    check(ws.cell(1, 1).value == 'id', 'header row: id')
    check(ws.cell(1, 2).value == 'name', 'header row: name')
    check(ws.cell(2, 1).value == 1, 'row 1 id')
    check(ws.cell(2, 2).value == 'alice', 'row 1 name')


def test_export_pdf():
    print("\n-- _export_pdf --")
    if not HAS_REPORTLAB:
        print("  SKIPPED (reportlab not installed)")
        return
    rows = [FakeRow(id=1, name='alice'), FakeRow(id=2, name='bob')]
    resp = _export_pdf('users', ['id', 'name'], rows, False)
    check(resp.status_code == 200, f'status 200 (got {resp.status_code})')
    check(resp.mimetype == 'application/pdf', 'correct pdf content-type')
    check(resp.data[:4] == b'%PDF', 'valid PDF header')


def test_export_csv():
    print("\n-- _export_csv --")
    rows = [FakeRow(id=1, name='alice')]
    resp = _export_csv('users', ['id', 'name'], rows, False)
    check(resp.mimetype == 'text/csv', 'content-type is text/csv')
    reader = csv.reader(io.StringIO(resp.data.decode()))
    headers = next(reader)
    check(headers == ['id', 'name'], 'CSV headers correct')
    data_row = next(reader)
    check(data_row == ['1', 'alice'], 'CSV data row correct')


def test_export_formats_registry():
    print("\n-- EXPORT_FORMATS registry --")
    check('json' in EXPORT_FORMATS, 'json registered')
    check('xlsx' in EXPORT_FORMATS, 'xlsx registered')
    check('pdf' in EXPORT_FORMATS, 'pdf registered')
    check(len(EXPORT_FORMATS) == 3, f'3 formats registered (got {len(EXPORT_FORMATS)})')


# ---------------------------------------------------------------------------
# Integration tests via Flask test client (needs DB)
# ---------------------------------------------------------------------------

def test_flask_export_routes():
    print("\n-- Flask export routes --")
    import tempfile

    tmp = tempfile.NamedTemporaryFile(suffix='.db', delete=False)
    tmp.close()
    db_path = tmp.name

    os.environ['DATABASE_URL'] = f'sqlite:///{db_path}'

    try:
        app = create_app()
        app.config['TESTING'] = True

        with app.test_client() as client:
            with app.app_context():
                from app.models import User, Module, QueryReport
                u1 = User.query.filter_by(username='export_tester').first()
                if not u1:
                    u1 = User(username='export_tester', password_hash='$2b$12$dummy',
                              role='admin', is_active=True, is_approved=True)
                    _db.session.add(u1)
                    _db.session.commit()

                mod = Module.query.filter_by(slug='test-export').first()
                if not mod:
                    mod = Module(name='Test Export', slug='test-export', version='1.0.0',
                                 author='test', enabled=True)
                    _db.session.add(mod)
                    _db.session.commit()

                q = QueryReport(
                    module_id=mod.id, name='Test Query',
                    sql='SELECT id, username FROM users LIMIT 5',
                    chart_type='none', label_column='', data_columns='',
                )
                _db.session.add(q)
                _db.session.commit()
                user_id = u1.id
                query_id = q.id

            # Log in via session
            with client.session_transaction() as sess:
                sess['_user_id'] = str(user_id)
                sess['_fresh'] = 'true'

            # Test query export route
            resp = client.get(f'/__admin/queries/{query_id}/export?format=json')
            check(resp.status_code == 200, f'query JSON export status {resp.status_code}')

            resp = client.get(f'/__admin/queries/{query_id}/export?format=xlsx')
            check(resp.status_code == 200, f'query XLSX export status {resp.status_code}')

            resp = client.get(f'/__admin/queries/{query_id}/export?format=pdf')
            check(resp.status_code == 200, f'query PDF export status {resp.status_code}')

            # Test unsupported format returns 400
            resp = client.get(f'/__admin/queries/{query_id}/export?format=xml')
            check(resp.status_code == 400, f'unsupported format returns 400 (got {resp.status_code})')

            # Test list_view dispatches all formats (via users route)
            resp = client.get('/__admin/users/', follow_redirects=True)
            check(resp.status_code == 200, f'users page loads (status {resp.status_code})')

            resp = client.get('/__admin/users?format=json', follow_redirects=True)
            check(resp.status_code == 200, f'users JSON export status {resp.status_code}')

            resp = client.get('/__admin/users?format=xlsx', follow_redirects=True)
            check(resp.status_code == 200, f'users XLSX export status {resp.status_code}')

            resp = client.get('/__admin/users?format=pdf', follow_redirects=True)
            check(resp.status_code == 200, f'users PDF export status {resp.status_code}')

            # CSV still works (regression)
            resp = client.get('/__admin/users?format=csv', follow_redirects=True)
            check(resp.status_code == 200, f'users CSV export still works (status {resp.status_code})')
            check(resp.mimetype == 'text/csv', 'CSV content-type preserved')

    finally:
        os.unlink(db_path)
        os.environ.pop('DATABASE_URL', None)


def main():
    print("Export Format Tests")
    print("=" * 50)

    print(f"openpyxl available: {HAS_OPENPYXL}")
    print(f"reportlab available: {HAS_REPORTLAB}")

    test_serialize_value()
    test_row_to_dict()
    test_export_json()
    test_export_xlsx()
    test_export_pdf()
    test_export_csv()
    test_export_formats_registry()
    test_flask_export_routes()

    print("\n" + "=" * 50)
    total = passed + failed
    print(f"Results: {passed}/{total} passed, {failed} failed")
    if failed:
        sys.exit(1)
    else:
        print("All tests passed!")


if __name__ == '__main__':
    main()
