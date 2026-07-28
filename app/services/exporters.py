"""Export utilities for JSON, Excel (xlsx), and PDF formats.

Each function takes (name_plural, columns, rows, has_module) and returns
a Flask Response with the appropriate Content-Type and Content-Disposition
headers for file download.

Rows are expected to be model instances or dicts with attribute-style access.
"""
import io
import json
from datetime import datetime, date, timezone
from decimal import Decimal

from flask import Response

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False


def _serialize_value(val):
    """Convert a value to a JSON-serializable type."""
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        return val.isoformat()
    if isinstance(val, Decimal):
        return float(val)
    if isinstance(val, bytes):
        return val.decode('utf-8', errors='replace')
    if hasattr(val, 'isoformat'):
        return str(val)
    return val


def _row_to_dict(row, columns, has_module=False):
    """Convert a row (model or dict) to a plain dict for export."""
    result = {}
    if has_module:
        mod = getattr(row, 'module', None)
        result['module'] = mod.name if mod else ''
    for col in columns:
        val = getattr(row, col, None) if not isinstance(row, dict) else row.get(col)
        result[col] = _serialize_value(val)
    return result


def _get_rows_data(name_plural, columns, rows, has_module):
    """Build a structured dict of export data from raw rows."""
    data = {
        'name': name_plural,
        'columns': list(columns),
        'has_module': has_module,
        'rows': [],
    }
    for r in rows:
        data['rows'].append(_row_to_dict(r, columns, has_module))
    return data


def _export_json(name_plural, columns, rows, has_module, metadata=None):
    """Export rows as JSON response."""
    data = _get_rows_data(name_plural, columns, rows, has_module)
    if metadata:
        data['metadata'] = metadata
    buf = io.StringIO()
    json.dump(data, buf, indent=2, default=str)
    filename = name_plural.replace(' ', '_') + '.json'
    return Response(
        buf.getvalue(),
        mimetype='application/json',
        headers={'Content-Disposition': f'attachment; filename={filename}'},
    )


def _export_xlsx(name_plural, columns, rows, has_module, metadata=None):
    """Export rows as an Excel (.xlsx) response."""
    if not HAS_OPENPYXL:
        return Response('Excel export requires openpyxl. Install with: pip install openpyxl',
                        status=501, mimetype='text/plain')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = name_plural[:31]

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    # Write metadata at top if provided
    meta_start_row = 1
    if metadata:
        ws.cell(row=1, column=1, value=name_plural).font = Font(bold=True, size=14)
        for i, (key, val) in enumerate(metadata.items(), 2):
            ws.cell(row=i, column=1, value=f'{key}:').font = Font(bold=True)
            ws.cell(row=i, column=2, value=str(val))
        meta_start_row = len(metadata) + 3

    headers = list(columns)
    if has_module:
        headers.insert(0, 'module')

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=meta_start_row, column=col_idx, value=str(header))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    for row_idx, row in enumerate(rows, meta_start_row + 1):
        vals = []
        if has_module:
            mod = getattr(row, 'module', None)
            vals.append(mod.name if mod else '')
        for col in columns:
            val = getattr(row, col, None) if not isinstance(row, dict) else row.get(col)
            vals.append(_serialize_value(val))
        for col_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border

    for col_idx in range(1, len(headers) + 1):
        max_len = 10
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max_len

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = name_plural.replace(' ', '_') + '.xlsx'
    return Response(
        buf.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename={filename}'},
    )


def _hex_to_reportlab_color(hex_color):
    """Convert hex color string to reportlab Color."""
    hex_color = hex_color.lstrip('#')
    r = int(hex_color[0:2], 16) / 255.0
    g = int(hex_color[2:4], 16) / 255.0
    b = int(hex_color[4:6], 16) / 255.0
    return colors.Color(r, g, b)


def _draw_bar_chart(canvas, chart_data, x, y, width, height):
    """Draw a basic bar chart on the PDF canvas."""
    if not chart_data or not chart_data.get('datasets'):
        return

    datasets = chart_data['datasets']
    labels = chart_data.get('labels', [])
    if not labels or not datasets[0]['data']:
        return

    # Find max value for scaling
    max_val = max(max(ds['data']) for ds in datasets) if datasets else 1
    if max_val == 0:
        max_val = 1

    # Chart area
    chart_left = x + 0.3 * inch
    chart_right = x + width - 0.3 * inch
    chart_bottom = y
    chart_top = y + height - 0.3 * inch
    chart_height = chart_top - chart_bottom
    chart_width = chart_right - chart_left

    # Draw axes
    canvas.setStrokeColor(colors.grey)
    canvas.setLineWidth(1)
    canvas.line(chart_left, chart_bottom, chart_left, chart_top)
    canvas.line(chart_left, chart_top, chart_right, chart_top)

    # Draw bars
    num_labels = len(labels)
    num_datasets = len(datasets)
    group_width = chart_width / num_labels if num_labels > 0 else 0
    bar_width = min(group_width * 0.6 / max(num_datasets, 1), 0.4 * inch)

    for i, label in enumerate(labels):
        group_x = chart_left + i * group_width + group_width / 2

        # Draw label
        canvas.setFont('Helvetica', 7)
        canvas.drawCentredString(group_x, chart_bottom - 0.15 * inch, str(label)[:15])

        # Draw bars for each dataset
        for j, ds in enumerate(datasets):
            val = ds['data'][i] if i < len(ds['data']) else 0
            bar_height = (val / max_val) * chart_height if max_val > 0 else 0
            bar_x = group_x - (num_datasets * bar_width) / 2 + j * bar_width
            bar_color = _hex_to_reportlab_color(ds.get('color', '#2563eb'))

            canvas.setFillColor(bar_color)
            canvas.setStrokeColor(colors.white)
            canvas.setLineWidth(0.5)
            canvas.rect(bar_x, chart_bottom, bar_width, bar_height, fill=1, stroke=1)

    # Draw title
    if chart_data.get('title'):
        canvas.setFont('Helvetica-Bold', 10)
        canvas.drawCentredString(x + width / 2, y + height - 0.2 * inch, chart_data['title'])


def _draw_pie_chart(canvas, chart_data, x, y, width, height):
    """Draw a basic pie chart on the PDF canvas."""
    if not chart_data or not chart_data.get('datasets'):
        return

    datasets = chart_data['datasets']
    labels = chart_data.get('labels', [])
    if not labels or not datasets[0]['data']:
        return

    data = datasets[0]['data']
    total = sum(data) if data else 1
    if total == 0:
        total = 1

    cx = x + width / 2
    cy = y + height / 2 - 0.2 * inch
    radius = min(width, height) / 4

    # Draw pie slices
    start_angle = 0
    for i, val in enumerate(data):
        if i >= len(labels):
            break
        slice_angle = (val / total) * 360
        color = _hex_to_reportlab_color(datasets[0].get('color', '#2563eb'))
        canvas.setFillColor(color)
        canvas.setStrokeColor(colors.white)
        canvas.setLineWidth(1)
        canvas.arc(cx - radius, cy - radius, cx + radius, cy + radius, start_angle, start_angle + slice_angle, fill=1, stroke=1)
        start_angle += slice_angle

    # Draw legend
    legend_x = x + width - 1.2 * inch
    legend_y = y + height - 0.5 * inch
    canvas.setFont('Helvetica', 8)
    for i, label in enumerate(labels[:8]):
        color = _hex_to_reportlab_color(datasets[0].get('color', '#2563eb'))
        canvas.setFillColor(color)
        canvas.rect(legend_x, legend_y - i * 0.2 * inch, 0.15 * inch, 0.15 * inch, fill=1)
        canvas.setFillColor(colors.black)
        canvas.drawString(legend_x + 0.2 * inch, legend_y - i * 0.2 * inch - 0.05 * inch, str(label)[:20])

    # Draw title
    if chart_data.get('title'):
        canvas.setFont('Helvetica-Bold', 10)
        canvas.drawCentredString(x + width / 2, y + height - 0.2 * inch, chart_data['title'])


def _export_pdf(name_plural, columns, rows, has_module, metadata=None, chart_data=None):
    """Export rows as a PDF response with formatted table and optional chart."""
    if not HAS_REPORTLAB:
        return Response('PDF export requires reportlab. Install with: pip install reportlab',
                        status=501, mimetype='text/plain')

    buf = io.BytesIO()

    # Calculate needed height: metadata + chart + table
    meta_height = 0
    if metadata:
        meta_height = len(metadata) * 0.15 * inch + 0.2 * inch

    chart_height = 0
    if chart_data and chart_data.get('datasets'):
        chart_height = 2.5 * inch

    table_height = min((len(rows) + 1) * 0.2 * inch, 4 * inch)

    total_height = meta_height + chart_height + table_height + 0.5 * inch
    page_height = max(letter[1], total_height + 1 * inch)

    pdf = SimpleDocTemplate(
        buf,
        pagesize=(letter[0], page_height),
        rightMargin=0.5 * inch,
        leftMargin=0.5 * inch,
        topMargin=0.5 * inch,
        bottomMargin=0.5 * inch,
    )

    elements = []

    # Title
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=6,
    )
    subtitle_style = ParagraphStyle(
        'Subtitle',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.grey,
        spaceAfter=12,
    )
    info_style = ParagraphStyle(
        'Info',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.HexColor('#555555'),
        spaceAfter=4,
    )
    elements.append(Paragraph(name_plural.title(), title_style))
    elements.append(Paragraph(
        f'{len(rows)} row{"s" if len(rows) != 1 else ""} exported',
        subtitle_style,
    ))

    # Metadata
    if metadata:
        elements.append(Spacer(1, 0.15 * inch))
        for key, val in metadata.items():
            elements.append(Paragraph(f'<b>{key}:</b> {val}', info_style))

    elements.append(Spacer(1, 0.25 * inch))

    # Draw chart using Canvas for more control
    if chart_data and chart_data.get('datasets'):
        from reportlab.pdfgen import canvas as pdf_canvas
        chart_canvas = pdf.Canvas(buf, pagesize=(letter[0], page_height))

        chart_type = chart_data.get('type', 'bar')
        chart_width = letter[0] - inch
        chart_x = 0.5 * inch
        chart_y = 0.5 * inch + table_height + 0.3 * inch

        if chart_type in ('pie', 'doughnut'):
            _draw_pie_chart(chart_canvas, chart_data, chart_x, chart_y, chart_width, chart_height)
        else:
            _draw_bar_chart(chart_canvas, chart_data, chart_x, chart_y, chart_width, chart_height)

        chart_canvas.showPage()
        chart_canvas.save()

    # Build table data
    headers = list(columns)
    if has_module:
        headers.insert(0, 'module')

    table_data = [headers]
    for row in rows:
        vals = []
        if has_module:
            mod = getattr(row, 'module', None)
            vals.append(mod.name if mod else '')
        for col in columns:
            val = getattr(row, col, None) if not isinstance(row, dict) else row.get(col)
            serialized = _serialize_value(val)
            vals.append(str(serialized) if serialized is not None else '')
        table_data.append(vals)

    # Build the PDF table
    if table_data:
        max_cols = len(table_data[0])
        col_widths = min(1.5 * inch, (letter[0] - inch) / max_cols)
        table = Table(table_data, repeatRows=1)

        header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid') if HAS_OPENPYXL else colors.HexColor('#2563EB')

        table_style = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563EB')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F3F4F6')]),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ]

        table.setStyle(TableStyle(table_style))
        elements.append(table)

    pdf.build(elements)
    buf.seek(0)
    filename = name_plural.replace(' ', '_') + '.pdf'
    return Response(
        buf.getvalue(),
        mimetype='application/pdf',
        headers={'Content-Disposition': f'attachment; filename={filename}'},
    )

    # Build table data
    headers = list(columns)
    if has_module:
        headers.insert(0, 'module')

    table_data = [headers]
    for row in rows:
        vals = []
        if has_module:
            mod = getattr(row, 'module', None)
            vals.append(mod.name if mod else '')
        for col in columns:
            val = getattr(row, col, None) if not isinstance(row, dict) else row.get(col)
            serialized = _serialize_value(val)
            vals.append(str(serialized) if serialized is not None else '')
        table_data.append(vals)

    # Build the PDF table
    if table_data:
        max_cols = len(table_data[0])
        col_widths = min(1.5 * inch, (letter[0] - inch) / max_cols)
        table = Table(table_data, repeatRows=1)

        header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid') if HAS_OPENPYXL else colors.HexColor('#2563EB')

        table_style = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563EB')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F3F4F6')]),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ]

        table.setStyle(TableStyle(table_style))
        elements.append(table)

    pdf.build(elements)
    buf.seek(0)
    filename = name_plural.replace(' ', '_') + '.pdf'
    return Response(
        buf.getvalue(),
        mimetype='application/pdf',
        headers={'Content-Disposition': f'attachment; filename={filename}'},
    )


EXPORT_FORMATS = {
    'json': _export_json,
    'xlsx': _export_xlsx,
    'pdf': _export_pdf,
}
